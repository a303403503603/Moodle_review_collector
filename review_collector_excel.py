from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def login(username, password,website):
    driver = webdriver.Chrome()
    driver.get(website)
    link = driver.find_element(by=By.CLASS_NAME, value="continuebutton")
    link.click()
    username_input = driver.find_element(by=By.ID, value="username")
    password_input = driver.find_element(by=By.ID, value="password")
    username_input.send_keys(username)
    password_input.send_keys(password)
    password_input.send_keys(Keys.RETURN)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID,"page-wrapper")))
    return driver

def search_first_page(driver,week):
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, "html.parser")
    container = (soup.find(class_="container-fluid")
                    .find(id="page-content")
                    .find(class_="course-content")
                    .find(attrs={"aria-label": week})
                    .find(class_="section img-text"))
    content = container.find_all(class_="activity questionnaire modtype_questionnaire")
    url_list = []
    for url in content:
        button = (url.find(class_="mod-indent-outer")
                        .find(class_="activityinstance")
                        .find(class_="dimmed"))
        link_href = button.get("href")
        url_list.append(link_href)
    return url_list

def search_next_page(driver, link):
    current_url = driver.current_url
    driver.get(link)
    time.sleep(2)
    activity_html_content = driver.page_source
    activity_soup = BeautifulSoup(activity_html_content, "html.parser")
    main_content = (activity_soup.find(id="page-wrapper")
                                .find(id="page")
                                .find(id="page-content")
                                .find(id="region-main-box")
                                .find(role="main")
                                .find(class_="mod_questionnaire_viewpage"))
    h2_text = main_content.find("h2").text
    all_responses_link = main_content.find(class_="allresponses")
    all_responses_url = all_responses_link.find("a")["href"]
    data = open_third_data(driver, h2_text, all_responses_url)
    driver.get(current_url)
    return h2_text, data

def open_third_data(driver, h2_text, all_responses_url):
    driver.get(all_responses_url)
    time.sleep(2)
    second_page_html_content = driver.page_source
    second_page_soup = BeautifulSoup(second_page_html_content, "html.parser")
    containers = (second_page_soup.find(id="page-wrapper")
                                    .find(id="page")
                                    .find(id="region-main")
                                    .find(role="main")
                                    .find(class_="box generalbox")
                                    .find_all(class_="qn-container"))
    last_container = containers[-1]
    feedback_table = (last_container.find(class_="qn-content")
                                    .find(class_="generaltable"))
    rows = (feedback_table.find("tbody")
                        .find_all("tr")[:-2])
    data = []
    for row in rows:
        reviewer_name = row.find("td",class_="cell c0").text.strip()
        reviewes = ""
        td_element = row.find("td",class_="cell c1 lastcol")
        for p_element in td_element.find_all("p"):
            reviewes += p_element.get_text(strip=True) + "\n"
        reviewes = reviewes.strip()
        if reviewes != '':
            data.append((reviewer_name, reviewes))
    return data
    
def save_to_excel(i, data, filename='c:\\Users\\a3034\\Desktop\\OUTPUT.xlsx'):
    if not os.path.isfile(filename):
        wb = Workbook()
        ws = wb.active
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active
    if i == 1:
        fill = PatternFill(start_color="9F4D95", end_color="FF00FF", fill_type="solid")
    elif i == 2:
        fill = PatternFill(start_color="FFD306", end_color="FFFF00", fill_type="solid")
    else:
        fills = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                 PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")]

    ws.column_dimensions['A'].width = 21.5 
    ws.column_dimensions['B'].width = 12 
    ws.column_dimensions['C'].width = 165 
    ws.column_dimensions['D'].width = 75  

    if i == 1 or i == 2:
        ws.append([data, '', '', ''])
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = Font(size=20, name='Microsoft YaHei UI', bold=True, italic=False)
            cell.alignment = Alignment(vertical='center')
            ws.row_dimensions[cell.row].height = 30

    else:
        for reviewer_name, review in data:
            ws.append(['', reviewer_name, review, ''])
            for cell in ws[ws.max_row]:
                cell.fill = fills[ws.max_row % 2]
                cell.font = Font(size=12, name='Microsoft YaHei UI', bold=False, italic=False)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                ws.row_dimensions[ws.max_row].auto_height = True

    wb.save(filename)

if __name__ == "__main__":
    website = "your URL"
    driver = login("your account number", "your password",website)
    week = input("Please enter the week number you are looking for: ")
    save_to_excel(1,week)
    print('login succeed')
    first_page_links = search_first_page(driver,week)
    for url in first_page_links:
        text,data = search_next_page(driver, url)
        save_to_excel(2,text)
        save_to_excel(3,data)
    driver.quit()
    print("end")
